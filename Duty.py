from Student import *
import pandas as pd 
from init import *
import openpyxl
import os



def arrange_duty(Students,week,day,lesson):
    best_candidate = None
    min_ordered = float('inf')

    for student in Students :
        if student.FreeTime[week][day][lesson]==1 :
            if student.Ordered < min_ordered :
                min_ordered=student.Ordered
                best_candidate = student
            elif student.Ordered== min_ordered :
                if Students.index(student)<Students.index(best_candidate):
                    best_candidate=student
 
    if best_candidate :
        best_candidate.Ordered+=1
        best_candidate.FreeTime[week][day][lesson]=0

        return best_candidate
    return None

def getTheans(Students):
    schedule = []
    for i in range(18):
        week_schedule = []
        for j in range(5):
            day_schedule = []
            for k in range(4):
                person=arrange_duty(Students,i,j,k)
                if person !=None :
                    name=person.name
                    depart=person.department
                    phone=person.phonenumber
                    value=name+"\n("+depart+")\n"+phone
                else :
                    value=None
                day_schedule.append(value)
            week_schedule.append(day_schedule)
        schedule.append(week_schedule)
    return schedule

def save_to_excel(schedule):

    folder_path = "值班表"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    # 遍历每个周，并生成单独的 Excel 文件
    for week in range(18):
        # 创建一个新的工作簿和工作表
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # 设置表头
        headers = ["课程", "周一", "周二", "周三", "周四", "周五"]
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = headers[col - 1]
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # 填充数据
        for lesson in range(4):
            row = lesson + 2  # 从第 2 行开始
            sheet.cell(row=row, column=1).value = f"第 {lesson + 1} 节课"
            
            for day in range(5):
                cell = sheet.cell(row=row, column=day + 2)
                cell.value = schedule[week][day][lesson]
        
        # 保存工作簿，文件名为“第X周值班表.xlsx”
        file_name = f"{folder_path}/第{week + 1}周值班表.xlsx"
        workbook.save(file_name)
        print(f"已生成第 {week + 1} 周值班表 - {file_name}")
        
    
    print("所有值班表已生成完毕！")

if __name__ == "__main__":
    Students=load_students_from_json("数据.json")
    schedule=WeeklySchedule()#创建 周几到数字的索引
    classes=Daytime()# 创建每节课到数字的索引
    schedule=getTheans(Students)
    save_to_excel(schedule)
    save_students_to_json(Students,'数据.json')
    